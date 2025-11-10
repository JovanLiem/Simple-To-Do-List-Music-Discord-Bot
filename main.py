import os
import asyncio
import asyncpg
import subprocess
import platform
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from discord.ext import commands, tasks
import discord
from dotenv import load_dotenv
import sys
import shutil
import yt_dlp
from collections import deque
from discord import app_commands
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import pytz

# =====================================================
# FFMPEG AUTO-INSTALLER (tetap sama)
# =====================================================
def check_ffmpeg():
    return shutil.which("ffmpeg") is not None

def install_ffmpeg():
    system = platform.system()
    if system != "Linux":
        print(f"Automatic installation only supports Linux. Current system: {system}")
        return False
    print("FFmpeg not found. Attempting to install...")
    try:
        with open("/etc/os-release", "r") as f:
            os_info = f.read().lower()
        if "ubuntu" in os_info or "debian" in os_info:
            subprocess.run(["sudo", "apt-get", "update"], check=True)
            subprocess.run(["sudo", "apt-get", "install", "-y", "ffmpeg"], check=True)
        elif "centos" in os_info or "rhel" in os_info or "fedora" in os_info:
            subprocess.run(["sudo", "yum", "install", "-y", "ffmpeg"], check=True)
        elif "arch" in os_info:
            subprocess.run(["sudo", "pacman", "-S", "--noconfirm", "ffmpeg"], check=True)
        else:
            print("Unknown Linux distribution. Install FFmpeg manually.")
            return False
        print("FFmpeg installed successfully!")
        return True
    except Exception as e:
        print(f"Error installing FFmpeg: {e}")
        return False

def ensure_ffmpeg():
    if check_ffmpeg():
        print("FFmpeg is already installed.")
        return True
    else:
        success = install_ffmpeg()
        if not success:
            print("Bot requires FFmpeg. Install manually and restart.")
            sys.exit(1)
        return True

ensure_ffmpeg()

# === Load env & setup ===
load_dotenv()
TOKEN = os.getenv("DISCORD_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")
FFMPEG_PATH = shutil.which("ffmpeg") or "ffmpeg"
WIB = ZoneInfo("Asia/Jakarta")

intents = discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix='!', intents=intents)

# GLOBAL QUEUE — PASTIKAN SELALU deque!
SONG_QUEUES = {}  # str(guild_id) -> deque

# =====================================================
# Database
# =====================================================
async def get_db():
    return await asyncpg.connect(DATABASE_URL)

async def init_db():
    conn = await asyncpg.connect(DATABASE_URL)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS todos (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            task_date DATE NOT NULL,
            task TEXT NOT NULL,
            done BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        );
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS music_history (
            id SERIAL PRIMARY KEY,
            guild_id BIGINT NOT NULL,
            user_id BIGINT NOT NULL,
            title TEXT NOT NULL,
            url TEXT,
            action TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW()
        );
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            username TEXT NOT NULL,
            guild_id BIGINT NOT NULL,
            checkin_time TIMESTAMPTZ DEFAULT NOW(),
            checkout_time TIMESTAMPTZ,
            work_duration INTERVAL
        );
    """)
    await conn.close()

@bot.event
async def on_ready():
    print(f"Logged in as {bot.user.name}")
    await init_db()
    try:
        synced = await bot.tree.sync()
        print(f"Synced {len(synced)} slash command(s)")
    except Exception as e:
        print(f"Failed to sync commands: {e}")

# =====================================================
# SAFE QUEUE HELPER
# =====================================================
def get_queue(guild_id: str):
    """Pastikan queue selalu deque, bahkan jika rusak."""
    if guild_id not in SONG_QUEUES:
        SONG_QUEUES[guild_id] = deque()
    queue = SONG_QUEUES[guild_id]
    # Jika entah bagaimana jadi coroutine (bug lama), perbaiki
    if asyncio.iscoroutine(queue):
        print(f"[FIX] Queue untuk {guild_id} rusak, diperbaiki.")
        SONG_QUEUES[guild_id] = deque()
    return SONG_QUEUES[guild_id]

# =====================================================
# Music Search Helper
# =====================================================
async def search_ytdlp_async(query, ydl_opts):
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, lambda: _extract(query, ydl_opts))

def _extract(query, ydl_opts):
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        return ydl.extract_info(query, download=False)

# =====================================================
# PLAY COMMAND
# =====================================================
@bot.tree.command(name="play", description="Putar lagu atau tambahkan ke antrean.")
@app_commands.describe(song_query="Judul lagu atau URL YouTube")
async def play(interaction: discord.Interaction, song_query: str):
    await interaction.response.defer()

    if not interaction.user.voice or not interaction.user.voice.channel:
        return await interaction.followup.send("Kamu harus berada di voice channel.")

    voice_channel = interaction.user.voice.channel
    voice_client = interaction.guild.voice_client

    if voice_client is None:
        voice_client = await voice_channel.connect()
    elif voice_channel != voice_client.channel:
        await voice_client.move_to(voice_channel)

    ydl_options = {"format": "bestaudio[abr<=96]/bestaudio", "noplaylist": True}
    query = "ytsearch1:" + song_query

    try:
        results = await search_ytdlp_async(query, ydl_options)
    except Exception as e:
        return await interaction.followup.send("Gagal mencari lagu. Coba lagi.")

    tracks = results.get("entries", [])
    if not tracks:
        return await interaction.followup.send("Lagu tidak ditemukan.")

    first_track = tracks[0]
    audio_url = first_track["url"]
    title = first_track.get("title", "Unknown Title")

    guild_id = str(interaction.guild_id)
    queue = get_queue(guild_id)  # PASTIKAN AMAN
    queue.append((audio_url, title))

    # Simpan ke DB
    conn = await get_db()
    await conn.execute(
        """INSERT INTO music_history (guild_id, user_id, title, url, action, created_at)
           VALUES ($1, $2, $3, $4, $5, $6)""",
        interaction.guild_id, interaction.user.id, title, audio_url,
        "queued" if voice_client.is_playing() else "played", datetime.now(WIB)
    )
    await conn.close()

    if voice_client.is_playing() or voice_client.is_paused():
        await interaction.followup.send(f"Ditambahkan ke antrean: **{title}**")
    else:
        await interaction.followup.send(f"Memutar sekarang: **{title}**")
        await play_next_song(voice_client, guild_id, interaction.channel)

@bot.event
async def on_voice_state_update(member: discord.Member, before: discord.VoiceState, after: discord.VoiceState):
    # Kalau bot stop playing → cek queue & play next ATAU disconnect
    if member.id == bot.user.id and before.channel and after.channel:
        vc = member.guild.voice_client
        if vc and not vc.is_playing() and not vc.is_paused():
            guild_id = str(member.guild.id)
            queue = get_queue(guild_id)
            if queue:
                # Ada queue → play next
                channel = discord.utils.get(member.guild.text_channels, name="general") or member.guild.text_channels[0]
                await play_next_song(vc, guild_id, channel)
            else:
                # Kosong → disconnect
                await asyncio.sleep(5)  # 5 detik grace period
                if vc and not vc.is_playing():
                    await vc.disconnect()       
                    
# =====================================================
# PLAY NEXT SONG (FIXED!)
# =====================================================
async def play_next_song(voice_client: discord.VoiceClient, guild_id: str, channel: discord.TextChannel):
    queue = get_queue(guild_id)
    
    if not queue:
        await channel.send("📭 Antrean selesai. Bot keluar dari VC.")
        if voice_client.is_connected():
            await voice_client.disconnect()  # ← INI YANG HILANG!
        return

    try:
        audio_url, title = queue.popleft()
        print(f"[MUSIC] Playing: {title}")
        
        ffmpeg_options = {
            "before_options": "-reconnect 1 -reconnect_streamed 1 -reconnect_delay_max 5 -reconnect_on_network_error 1",
            "options": "-vn -c:a libopus -b:a 96k",
        }
        
        source = discord.FFmpegOpusAudio(audio_url, **ffmpeg_options, executable=FFMPEG_PATH)
        
        # NON-BLOCKING after callback
        def after_play(error):
            if error:
                print(f"[ERROR] Playback failed: {error}")
            # JANGAN pakai .result() → langsung schedule
            asyncio.run_coroutine_threadsafe(
                play_next_song(voice_client, guild_id, channel), 
                bot.loop
            )
        
        voice_client.play(source, after=after_play)
        await channel.send(f"🎵 **Sekarang memutar: {title}**")
        
    except Exception as e:
        print(f"[CRITICAL] Play failed: {e}")
        await channel.send("❌ Gagal memutar lagu. Skip ke next.")
        await play_next_song(voice_client, guild_id, channel)  # Recursive skip


# =====================================================
# STOP, HISTORY, DLL (tetap sama, tapi aman)
# =====================================================
@bot.tree.command(name="stop", description="Hentikan musik dan disconnect.")
async def stop(interaction: discord.Interaction):
    voice_client = interaction.guild.voice_client
    if not voice_client:
        return await interaction.response.send_message("Bot tidak di voice channel.")

    guild_id = str(interaction.guild_id)
    get_queue(guild_id).clear()

    if voice_client.is_playing():
        voice_client.stop()
    await voice_client.disconnect()
    await interaction.response.send_message("Musik dihentikan dan bot keluar.")

@bot.tree.command(name="history", description="Lihat riwayat musik server ini.")
async def history(interaction: discord.Interaction):
    conn = await get_db()
    rows = await conn.fetch(
        """
        SELECT title, action, created_at AT TIME ZONE 'Asia/Jakarta' AS waktu
        FROM music_history
        WHERE guild_id = $1
        ORDER BY created_at DESC
        LIMIT 10;
        """,
        interaction.guild_id
    )
    await conn.close()

    if not rows:
        return await interaction.response.send_message("📭 Belum ada lagu yang pernah diputar di server ini.")

    msg_lines = ["🎧 **Riwayat 10 Lagu Terakhir:**\n"]

    for r in rows:
        waktu = r["waktu"].strftime("%Y-%m-%d %H:%M:%S")
        icon = "▶️" if r["action"] == "played" else "➕"
        # batasi panjang judul agar tidak pecah di HP
        title = r["title"]
        if len(title) > 60:
            title = title[:57] + "..."

        msg_lines.append(
            f"━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"{icon} **Lagu:** {title}\n"
            f"📀 **Status:** {r['action'].capitalize()}\n"
            f"🕒 **Waktu:** {waktu} WIB"
        )

    msg_text = "\n".join(msg_lines)
    await interaction.response.send_message(msg_text)

@bot.tree.command(name="next", description="Skip lagu sekarang dan putar lagu berikutnya.")
async def next(interaction: discord.Interaction):
    voice_client = interaction.guild.voice_client
    guild_id = str(interaction.guild_id)
    channel = interaction.channel

    if not voice_client or not voice_client.is_connected():
        return await interaction.response.send_message("❌ Bot tidak sedang di voice channel.", ephemeral=True)

    queue = get_queue(guild_id)
    if not queue:
        return await interaction.response.send_message("📭 Tidak ada lagu berikutnya dalam antrean.", ephemeral=True)

    # Stop current song → after callback akan auto-trigger next
    if voice_client.is_playing():
        voice_client.stop()
        await interaction.response.send_message("⏭️ Lagu dilewati, memutar lagu berikutnya...")
    else:
        await interaction.response.send_message("⏭️ Tidak sedang memutar lagu, mencoba lanjut ke berikutnya...")
        await play_next_song(voice_client, guild_id, channel)


@bot.tree.command(name="add", description="Tambah tugas ke daftar to-do kamu.")
@app_commands.describe(date_str="Tanggal (YYYY-MM-DD, opsional)", task="Deskripsi tugas")
async def add(interaction: discord.Interaction, task: str, date_str: str = ""):
    user_id = interaction.user.id
    now = datetime.now(WIB)

    # parsing tanggal
    try:
        task_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else now.date()
    except ValueError:
        return await interaction.response.send_message("⚠️ Format tanggal salah. Gunakan YYYY-MM-DD.", ephemeral=True)

    conn = await asyncpg.connect(DATABASE_URL)
    await conn.execute(
        "INSERT INTO todos (user_id, task_date, task, done, created_at) VALUES ($1, $2, $3, FALSE, $4);",
        user_id, task_date, task, now
    )
    await conn.close()
    await interaction.response.send_message(f"📝 Ditambahkan: **{task}** untuk **{task_date}**")

@bot.tree.command(name="list", description="Tampilkan daftar tugas kamu.")
@app_commands.describe(date_str="Tanggal (YYYY-MM-DD, opsional)")
async def list_tasks(interaction: discord.Interaction, date_str: str = ""):
    user_id = interaction.user.id
    now = datetime.now(WIB)
    today = now.date()

    # parsing tanggal
    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else today
    except ValueError:
        return await interaction.response.send_message("⚠️ Format tanggal tidak valid.", ephemeral=True)

    conn = await asyncpg.connect(DATABASE_URL)
    rows = await conn.fetch(
        "SELECT id, task, done FROM todos WHERE user_id=$1 AND task_date=$2 ORDER BY id;",
        user_id, target_date
    )
    await conn.close()

    if not rows:
        return await interaction.response.send_message(f"✨ Tidak ada tugas untuk **{target_date}**.")

    msg = [f"📅 **Tugas untuk {target_date}:**"]
    for row in rows:
        status = "✅" if row["done"] else "☐"
        msg.append(f"{status} {row['task']} (ID: {row['id']})")

    await interaction.response.send_message("\n".join(msg))

@bot.tree.command(name="done", description="Tandai tugas sebagai selesai.")
@app_commands.describe(task_id="ID tugas yang ingin ditandai selesai")
async def done(interaction: discord.Interaction, task_id: int):
    user_id = interaction.user.id
    conn = await asyncpg.connect(DATABASE_URL)
    result = await conn.execute("UPDATE todos SET done=TRUE WHERE id=$1 AND user_id=$2;", task_id, user_id)
    await conn.close()

    if result == "UPDATE 1":
        await interaction.response.send_message(f"✅ Tugas dengan ID {task_id} telah selesai!")
    else:
        await interaction.response.send_message("❌ ID tugas tidak ditemukan.")

@bot.tree.command(name="delete", description="Hapus tugas berdasarkan ID.")
@app_commands.describe(task_id="ID tugas yang ingin dihapus")
async def delete(interaction: discord.Interaction, task_id: int):
    user_id = interaction.user.id
    conn = await asyncpg.connect(DATABASE_URL)
    result = await conn.execute("DELETE FROM todos WHERE id=$1 AND user_id=$2;", task_id, user_id)
    await conn.close()

    if result == "DELETE 1":
        await interaction.response.send_message(f"🗑️ Tugas dengan ID {task_id} telah dihapus.")
    else:
        await interaction.response.send_message("❌ ID tugas tidak ditemukan.")

@bot.tree.command(name="clear", description="Hapus semua tugas untuk tanggal tertentu (default: hari ini).")
@app_commands.describe(date_str="Tanggal (YYYY-MM-DD, opsional)")
async def clear(interaction: discord.Interaction, date_str: str = ""):
    user_id = interaction.user.id
    today = datetime.now(WIB).date()

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else today
    except ValueError:
        return await interaction.response.send_message("⚠️ Format tanggal tidak valid.", ephemeral=True)

    conn = await asyncpg.connect(DATABASE_URL)
    result = await conn.execute("DELETE FROM todos WHERE user_id=$1 AND task_date=$2;", user_id, target_date)
    await conn.close()

    await interaction.response.send_message(f"🧹 Semua tugas untuk {target_date} telah dihapus.")

@bot.tree.command(name="dates", description="Lihat semua tugas kamu, dikelompokkan per tanggal.")
async def dates(interaction: discord.Interaction):
    user_id = interaction.user.id
    conn = await asyncpg.connect(DATABASE_URL)
    rows = await conn.fetch(
        "SELECT task_date, task, done FROM todos WHERE user_id=$1 ORDER BY task_date ASC, id ASC;",
        user_id
    )
    await conn.close()

    if not rows:
        return await interaction.response.send_message("✨ Kamu belum memiliki tugas sama sekali.")

    grouped = {}
    for r in rows:
        date_str = r["task_date"].strftime("%Y-%m-%d")
        grouped.setdefault(date_str, []).append(r)

    messages = ["📅 **Daftar Semua Tugas (WIB):**"]
    for date_str, tasks in grouped.items():
        messages.append(f"\n📆 {date_str}:")
        for t in tasks:
            status = "✅" if t["done"] else "☐"
            messages.append(f"　{status} {t['task']}")

    # pastikan tidak lebih dari 2000 karakter
    final_msg = ""
    for line in messages:
        if len(final_msg) + len(line) + 1 > 1900:
            await interaction.followup.send(final_msg)
            final_msg = ""
        final_msg += line + "\n"

    if final_msg:
        await interaction.response.send_message(final_msg)
        
@bot.tree.command(name="export_excel", description="Ekspor tugas kamu ke file Excel (bisa filter tanggal).")
@app_commands.describe(
    start_date="Tanggal mulai (YYYY-MM-DD, opsional)",
    end_date="Tanggal akhir (YYYY-MM-DD, opsional)"
)
async def export_excel(interaction: discord.Interaction, start_date: str = "", end_date: str = ""):
   	# Waktu WIB manual (tanpa pytz)
    def now_wib():
        return datetime.now(WIB)

    user_id = interaction.user.id
    user_name = interaction.user.name
    await interaction.response.defer(thinking=True)

    # Parsing tanggal
    date_filter = ""
    params = [user_id]
    try:
        if start_date and end_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
            date_filter = "AND task_date BETWEEN $2 AND $3"
            params.extend([start_dt, end_dt])
        elif start_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            date_filter = "AND task_date >= $2"
            params.append(start_dt)
        elif end_date:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
            date_filter = "AND task_date <= $2"
            params.append(end_dt)
    except ValueError:
        return await interaction.followup.send("⚠️ Format tanggal salah. Gunakan format `YYYY-MM-DD`.", ephemeral=True)

    # Query data
    conn = await asyncpg.connect(DATABASE_URL)
    query = f"""
        SELECT task_date, task, done, created_at AT TIME ZONE 'Asia/Jakarta' AS waktu_buat
        FROM todos
        WHERE user_id=$1 {date_filter}
        ORDER BY task_date ASC, id ASC;
    """
    rows = await conn.fetch(query, *params)
    await conn.close()

    if not rows:
        return await interaction.followup.send("📭 Tidak ada tugas dalam rentang tanggal tersebut.")

    # Buat workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Tugas"

    # Header
    ws.append(["Tanggal", "Deskripsi Tugas", "Status", "Dibuat Pada"])

    # Tambahkan isi data
    for r in rows:
        tanggal = r["task_date"].strftime("%Y-%m-%d")
        status = "✅ Selesai" if r["done"] else "☐ Belum"
        dibuat = r["waktu_buat"].strftime("%Y-%m-%d %H:%M:%S")
        ws.append([tanggal, r["task"], status, dibuat])

    # Gaya border
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Semua sel diberi border
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    # Gabungkan cell tanggal yang sama
    current_date = None
    start_row = None
    for i in range(2, ws.max_row + 1):
        tanggal = ws.cell(i, 1).value
        if tanggal != current_date:
            if start_row is not None and i - start_row > 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
            current_date = tanggal
            start_row = i
    # Merge blok terakhir
    if start_row is not None and ws.max_row - start_row >= 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)

    # Auto lebar kolom
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Simpan ke buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today_str = now_wib().strftime("%Y-%m-%d")
    filename = f"todo_{user_name}_{today_str}.xlsx"
    file = discord.File(output, filename=filename)
    await interaction.followup.send("📂 Berikut file Excel tugas kamu:", file=file)
    
@bot.tree.command(name="checkin", description="Catat absensi harian kamu (check-in).")
async def checkin(interaction: discord.Interaction):
    conn = await asyncpg.connect(DATABASE_URL)

    user_id = interaction.user.id
    username = interaction.user.name
    guild_id = interaction.guild_id
    wib = pytz.timezone("Asia/Jakarta")
    now_wib = datetime.now(wib)

    # Cek apakah user sudah check-in hari ini
    record = await conn.fetchrow("""
        SELECT id FROM attendance
        WHERE user_id = $1 AND guild_id = $2
        AND DATE(checkin_time AT TIME ZONE 'Asia/Jakarta') = CURRENT_DATE
    """, user_id, guild_id)

    if record:
        await interaction.response.send_message("⚠️ Kamu sudah check-in hari ini!")
        await conn.close()
        return

    # Simpan check-in baru
    await conn.execute("""
        INSERT INTO attendance (user_id, username, guild_id, checkin_time)
        VALUES ($1, $2, $3, $4)
    """, user_id, username, guild_id, now_wib)

    await conn.close()

    await interaction.response.send_message(
        f"✅ {username}, kamu berhasil check-in pada **{now_wib.strftime('%Y-%m-%d %H:%M:%S')} WIB**!"
    )

@bot.tree.command(name="checkout", description="Catat waktu pulang kamu (checkout).")
async def checkout(interaction: discord.Interaction):
    conn = await asyncpg.connect(DATABASE_URL)

    user_id = interaction.user.id
    guild_id = interaction.guild_id
    wib = pytz.timezone("Asia/Jakarta")
    now_wib = datetime.now(wib)

    # Ambil data checkin hari ini
    record = await conn.fetchrow("""
        SELECT id, checkin_time, checkout_time
        FROM attendance
        WHERE user_id = $1 AND guild_id = $2
        AND DATE(checkin_time AT TIME ZONE 'Asia/Jakarta') = CURRENT_DATE
        ORDER BY checkin_time DESC LIMIT 1
    """, user_id, guild_id)

    if not record:
        await interaction.response.send_message("⚠️ Kamu belum check-in hari ini.")
        await conn.close()
        return

    if record["checkout_time"]:
        await interaction.response.send_message("🕓 Kamu sudah checkout hari ini.")
        await conn.close()
        return

    checkin_time = record["checkin_time"].astimezone(wib)
    work_duration = now_wib - checkin_time

    # Update checkout_time dan durasi kerja
    await conn.execute("""
        UPDATE attendance
        SET checkout_time = $1, work_duration = $2
        WHERE id = $3
    """, now_wib, work_duration, record["id"])

    await conn.close()

    hours, remainder = divmod(work_duration.total_seconds(), 3600)
    minutes, _ = divmod(remainder, 60)

    await interaction.response.send_message(
        f"👋 Checkout berhasil pada **{now_wib.strftime('%Y-%m-%d %H:%M:%S')} WIB**!\n"
        f"⏰ Durasi kerja hari ini: **{int(hours)} jam {int(minutes)} menit.**"
    )
    
@bot.tree.command(name="riwayat_absensi", description="Lihat riwayat absensi kamu (5 hari terakhir).")
async def riwayat_absensi(interaction: discord.Interaction):
    conn = await asyncpg.connect(DATABASE_URL)
    user_id = interaction.user.id
    wib = pytz.timezone("Asia/Jakarta")

    rows = await conn.fetch("""
        SELECT 
            checkin_time AT TIME ZONE 'Asia/Jakarta' AS checkin,
            checkout_time AT TIME ZONE 'Asia/Jakarta' AS checkout,
            work_duration
        FROM attendance
        WHERE user_id = $1
        ORDER BY checkin_time DESC
        LIMIT 5
    """, user_id)
    await conn.close()

    if not rows:
        await interaction.response.send_message("📭 Kamu belum punya riwayat absensi.")
        return

    msg = "**🗓️ Riwayat Absensi Terakhir:**\n"
    for r in rows:
        checkin_str = r["checkin"].strftime("%Y-%m-%d %H:%M:%S") if r["checkin"] else "-"
        checkout_str = r["checkout"].strftime("%Y-%m-%d %H:%M:%S") if r["checkout"] else "-"
        durasi = str(r["work_duration"]).split(".")[0] if r["work_duration"] else "-"
        msg += f"📅 {checkin_str} → {checkout_str} | ⏱️ {durasi}\n"

    await interaction.response.send_message(msg)

@bot.command()
@commands.is_owner()
async def restart(ctx):
    await ctx.send("Bot akan restart...")
    await bot.close()
    os.execv(sys.executable, ['python'] + sys.argv)

# =====================================================
# RUN BOT
# =====================================================
if __name__ == "__main__":
    bot.run(TOKEN)